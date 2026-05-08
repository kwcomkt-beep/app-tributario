"""
=============================================================
  PIS/COFINS PRO  |  v8.0  – correcao critica de bugs
  Recuperacao Tributaria · Simples Nacional · Regime Revenda
=============================================================
  CORRECOES v8:
  - Leitura de data (dhEmi) robusta com namespace
  - Pipeline sem cache em bytes (evita corrupcao silenciosa)
  - Desserializacao correta do PGDAS (tipos numericos)
  - Calculo de credito corrigido
  - Segregacao corretamente avaliada
  - Plotly opcional com fallback nativo Streamlit
  - Classificacao NCM verificada
=============================================================
  Dependencias minimas: pip install streamlit pandas openpyxl
  Opcionais:           pip install plotly reportlab
=============================================================
"""
from __future__ import annotations

import io, logging, os, zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from io import BytesIO
from typing import Optional

import pandas as pd
import streamlit as st

# ── bibliotecas opcionais ─────────────────────────────────────
try:
    import plotly.express as px
    import plotly.graph_objects as go
    PLOTLY_OK = True
except ImportError:
    PLOTLY_OK = False

try:
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ─────────────────────────────────────────────────────────────
#  CONSTANTES
# ─────────────────────────────────────────────────────────────
VERSAO           = "8.0"
UPLOAD_MAX_MB    = 50
DIVERGENCIA_MAX  = 0.05
ALIQ_ESTIMATIVA  = 0.0925
SEG_PARCIAL_MIN  = 0.05
SEG_TOTAL_MIN    = 0.90

CST_MONO    = {"04","4"}
CST_NORMAL  = {"01","1","02","2","05","5","06","6","07","7","08","8","09","9"}
CFOP_REV    = {"5102","5403","5405","6102","6403","6404","5101","6101","5104","6104"}
CFOP_INC    = {"5201","5202","6201","6202","5910","6910","5949","6949"}

TABELA_NCM: dict[str, str] = {
    "27101112":"Gasolina automotiva comum",   "27101113":"Gasolina automotiva premium",
    "27101121":"Querosene de aviacao",         "27101131":"Oleo diesel",
    "27101500":"Oleos lubrificantes",          "27111100":"GNL",
    "27111910":"GLP",                          "30011000":"Glandulas/orgaos opoterapicos",
    "30021000":"Antissoros/imunoglobulinas",   "30022000":"Vacinas medicina humana",
    "30023000":"Vacinas medicina veterinaria", "30031000":"Medicamentos c/ penicilinas",
    "30032000":"Medicamentos c/ antibioticos", "30039099":"Outros medicamentos mistura",
    "30041000":"Med. penicilinas (doses)",     "30042000":"Med. antibioticos (doses)",
    "30043900":"Med. hormonais",               "30049099":"Outros med. uso humano",
    "33011000":"Oleos ess. citricos",          "33012900":"Outros oleos essenciais",
    "33030010":"Perfumes (extratos)",          "33030020":"Aguas-de-colonia",
    "33041000":"Maquiagem labios",             "33042000":"Sombras e delineadores",
    "33049900":"Outros prod. beleza",          "33051000":"Xampus",
    "33052000":"Preparacoes ondulacao",        "33053000":"Laques",
    "33059000":"Outras prep. capilares",       "33061000":"Dentifricio",
    "33062000":"Fio dental",                   "33069000":"Higiene bucal outros",
    "33071000":"Prep. barbear",                "33072000":"Desodorantes/antiperspirantes",
    "33074900":"Outros toucador",              "22011000":"Agua mineral/gaseificada",
    "22019000":"Outras aguas",                 "22021000":"Agua c/ acucar/adocante",
    "22029000":"Bebidas nao alcoolicas",       "22030000":"Cerveja de malte",
    "22060000":"Bebidas fermentadas",          "22071000":"Alcool etilico >= 80%",
    "22082000":"Conhaque",                     "22083000":"Uisque",
    "22084000":"Rum e tafia",                  "22085000":"Gim e genebra",
    "22086000":"Vodca",                        "22087000":"Licores",
    "22089900":"Outras beb. alcoolicas",       "87031000":"Veiculos neve/quadriciclos",
    "87032100":"Automoveis <= 1000 cm3",       "87032200":"Automoveis 1000-1500 cm3",
    "87032300":"Automoveis 1500-3000 cm3",     "87032400":"Automoveis > 3000 cm3",
    "87033300":"Automoveis diesel > 2500 cm3", "87060010":"Chassis c/ motor",
    "87089900":"Acessorios veiculos",          "87111000":"Motos <= 50 cm3",
    "87112000":"Motos 50-250 cm3",             "87113000":"Motos 250-500 cm3",
    "87114000":"Motos 500-800 cm3",            "87115000":"Motos > 800 cm3",
    "40111000":"Pneus novos automoveis",       "40112000":"Pneus novos onibus/caminhoes",
    "40113000":"Pneus novos avioes",           "40114000":"Pneus novos motocicletas",
    "40119100":"Pneus novos outros",           "40121100":"Pneus recauch. automoveis",
    "40121200":"Pneus recauch. onibus/caminhoes",
}

TABELAS_SIMPLES: dict[str, list[dict]] = {
    "Anexo I – Comercio": [
        {"faixa":1,"limite":180_000,   "aliquota":0.04,  "deducao":0.0},
        {"faixa":2,"limite":360_000,   "aliquota":0.073, "deducao":5_940.0},
        {"faixa":3,"limite":720_000,   "aliquota":0.095, "deducao":13_860.0},
        {"faixa":4,"limite":1_800_000, "aliquota":0.107, "deducao":22_500.0},
        {"faixa":5,"limite":3_600_000, "aliquota":0.143, "deducao":87_300.0},
        {"faixa":6,"limite":4_800_000, "aliquota":0.19,  "deducao":378_000.0},
    ],
    "Anexo II – Industria": [
        {"faixa":1,"limite":180_000,   "aliquota":0.045, "deducao":0.0},
        {"faixa":2,"limite":360_000,   "aliquota":0.078, "deducao":5_940.0},
        {"faixa":3,"limite":720_000,   "aliquota":0.10,  "deducao":13_860.0},
        {"faixa":4,"limite":1_800_000, "aliquota":0.113, "deducao":22_500.0},
        {"faixa":5,"limite":3_600_000, "aliquota":0.147, "deducao":85_500.0},
        {"faixa":6,"limite":4_800_000, "aliquota":0.30,  "deducao":720_000.0},
    ],
    "Anexo III – Servicos A": [
        {"faixa":1,"limite":180_000,   "aliquota":0.06,  "deducao":0.0},
        {"faixa":2,"limite":360_000,   "aliquota":0.112, "deducao":9_360.0},
        {"faixa":3,"limite":720_000,   "aliquota":0.135, "deducao":17_640.0},
        {"faixa":4,"limite":1_800_000, "aliquota":0.16,  "deducao":35_640.0},
        {"faixa":5,"limite":3_600_000, "aliquota":0.21,  "deducao":125_640.0},
        {"faixa":6,"limite":4_800_000, "aliquota":0.33,  "deducao":648_000.0},
    ],
    "Anexo IV – Servicos B": [
        {"faixa":1,"limite":180_000,   "aliquota":0.045, "deducao":0.0},
        {"faixa":2,"limite":360_000,   "aliquota":0.09,  "deducao":8_100.0},
        {"faixa":3,"limite":720_000,   "aliquota":0.102, "deducao":12_420.0},
        {"faixa":4,"limite":1_800_000, "aliquota":0.14,  "deducao":39_780.0},
        {"faixa":5,"limite":3_600_000, "aliquota":0.22,  "deducao":183_780.0},
        {"faixa":6,"limite":4_800_000, "aliquota":0.33,  "deducao":828_000.0},
    ],
    "Anexo V – Servicos C": [
        {"faixa":1,"limite":180_000,   "aliquota":0.15,  "deducao":0.0},
        {"faixa":2,"limite":360_000,   "aliquota":0.18,  "deducao":5_400.0},
        {"faixa":3,"limite":720_000,   "aliquota":0.195, "deducao":13_500.0},
        {"faixa":4,"limite":1_800_000, "aliquota":0.205, "deducao":20_700.0},
        {"faixa":5,"limite":3_600_000, "aliquota":0.23,  "deducao":62_100.0},
        {"faixa":6,"limite":4_800_000, "aliquota":0.305, "deducao":540_000.0},
    ],
}

REPARTICAO: dict[str, dict[int, dict]] = {
    "Anexo I – Comercio":     {1:{"pis":0.0,"cofins":0.0},    2:{"pis":0.0276,"cofins":0.1274},3:{"pis":0.0276,"cofins":0.1274},4:{"pis":0.0276,"cofins":0.1274},5:{"pis":0.0276,"cofins":0.1274},6:{"pis":0.0276,"cofins":0.1274}},
    "Anexo II – Industria":   {1:{"pis":0.0,"cofins":0.0},    2:{"pis":0.0186,"cofins":0.086}, 3:{"pis":0.0186,"cofins":0.086}, 4:{"pis":0.0186,"cofins":0.086}, 5:{"pis":0.0186,"cofins":0.086}, 6:{"pis":0.0186,"cofins":0.086}},
    "Anexo III – Servicos A": {1:{"pis":0.0,"cofins":0.0},    2:{"pis":0.0167,"cofins":0.0773},3:{"pis":0.0167,"cofins":0.0773},4:{"pis":0.0167,"cofins":0.0773},5:{"pis":0.0167,"cofins":0.0773},6:{"pis":0.0167,"cofins":0.0773}},
    "Anexo IV – Servicos B":  {1:{"pis":0.0,"cofins":0.0},    2:{"pis":0.0167,"cofins":0.0773},3:{"pis":0.0167,"cofins":0.0773},4:{"pis":0.0167,"cofins":0.0773},5:{"pis":0.0167,"cofins":0.0773},6:{"pis":0.0167,"cofins":0.0773}},
    "Anexo V – Servicos C":   {1:{"pis":0.0,"cofins":0.0},    2:{"pis":0.0098,"cofins":0.0454},3:{"pis":0.0098,"cofins":0.0454},4:{"pis":0.0098,"cofins":0.0454},5:{"pis":0.0098,"cofins":0.0454},6:{"pis":0.0098,"cofins":0.0454}},
}

# ─────────────────────────────────────────────────────────────
#  UTILS
# ─────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger("piscofins_pro")
_LOGS: list[dict] = []

def log_a(nivel: str, cat: str, msg: str, det: str = "") -> None:
    _LOGS.append({"ts": datetime.now().strftime("%H:%M:%S"),
                  "nivel": nivel, "categoria": cat,
                  "mensagem": msg, "detalhe": det})

def get_logs() -> list[dict]: return list(_LOGS)
def clear_logs() -> None: _LOGS.clear()

def r2(v: Optional[float]) -> Optional[float]:
    return round(v, 2) if v is not None else None

def brl(v: Optional[float]) -> str:
    if v is None: return "–"
    return "R$ {:,.2f}".format(v).replace(",","X").replace(".",",").replace("X",".")

def pct_s(v: Optional[float], casas: int = 2) -> str:
    if v is None: return "–"
    return "{:.{}f}%".format(v * 100, casas).replace(".", ",")

def fmt_mes(m: str) -> str:
    try: return datetime.strptime(m, "%Y-%m").strftime("%b/%Y")
    except: return m

def safe_float(v, default: float = 0.0) -> float:
    """Converte qualquer valor para float de forma segura."""
    if v is None: return default
    try: return float(str(v).replace(",", ".").replace("R$", "").strip())
    except: return default


# ─────────────────────────────────────────────────────────────
#  XML PARSER  — leitura robusta (sem cache de bytes)
# ─────────────────────────────────────────────────────────────
def _ltag(node) -> str:
    t = node.tag
    return t.split("}")[-1] if "}" in t else t

def _find(el, tag: str):
    """Busca tag ignorando namespace em qualquer profundidade."""
    for n in el.iter():
        if _ltag(n) == tag:
            return n
    return None

def _txt(el, tag: str, default: str = "") -> str:
    n = _find(el, tag)
    return n.text.strip() if n is not None and n.text else default

def _extrair_mes(root) -> str:
    """
    Extrai mes/ano da NF-e.
    Tenta dhEmi primeiro, depois dEmi.
    Aceita formatos: 2025-01-15T10:30:00-03:00, 2025-01-15, 2025-01
    """
    dh = _txt(root, "dhEmi") or _txt(root, "dEmi")
    if not dh:
        log_a("AVISO", "XMLParser", "dhEmi/dEmi ausente no XML")
        return "SEM-DATA"
    # Normaliza: pega apenas os primeiros 7 chars "YYYY-MM"
    try:
        return dh.strip()[:7]
    except Exception:
        return "SEM-DATA"

def _extrair_cnpj(root) -> str:
    # Tenta emitente primeiro
    emit = _find(root, "emit")
    if emit is not None:
        cnpj = _txt(emit, "CNPJ")
        if cnpj: return cnpj
    return _txt(root, "CNPJ") or ""

def ler_xml_nfe(nome: str, conteudo: bytes) -> tuple[list[dict], str, str]:
    """
    Le NF-e XML sem cache (cache de bytes causa corrupcao no Streamlit).
    Retorna (itens, mes_ano, cnpj).
    """
    raw = conteudo.lstrip(b"\xef\xbb\xbf").strip()
    if len(raw) / 1_048_576 > UPLOAD_MAX_MB:
        raise ValueError("Arquivo excede {}MB".format(UPLOAD_MAX_MB))
    try:
        root = ET.fromstring(raw)
    except ET.ParseError as e:
        raise ValueError("XML malformado: {}".format(e))

    mes  = _extrair_mes(root)
    cnpj = _extrair_cnpj(root)

    dets = [n for n in root.iter() if _ltag(n) == "det"]
    if not dets:
        log_a("AVISO", "XMLParser", "Nenhum <det> encontrado", nome)

    itens: list[dict] = []
    for det in dets:
        prod = _find(det, "prod")
        if not prod:
            continue

        ncm_raw = _txt(prod, "NCM")
        vprod   = _txt(prod, "vProd")
        cfop    = _txt(prod, "CFOP")

        # CST PIS
        cst_p = ""
        pis_node = _find(det, "PIS")
        if pis_node is not None:
            cst_p = _txt(pis_node, "CST")

        # CST COFINS
        cst_c = ""
        cof_node = _find(det, "COFINS")
        if cof_node is not None:
            cst_c = _txt(cof_node, "CST")

        # Valor
        try:
            valor = round(float(vprod.replace(",", ".")), 2)
        except ValueError:
            valor = 0.0
            log_a("AVISO", "XMLParser", "vProd invalido em {}".format(nome), ncm_raw)

        # NCM limpo (apenas digitos)
        ncm_limpo = "".join(c for c in ncm_raw if c.isdigit())

        itens.append({
            "descricao":  _txt(prod, "xProd") or "(sem desc.)",
            "ncm_raw":    ncm_raw,
            "ncm":        ncm_limpo,
            "valor":      valor,
            "cfop":       cfop,
            "cst_pis":    cst_p,
            "cst_cofins": cst_c,
        })

    log_a("INFO", "XMLParser",
          "Mes={} NCMs={} Itens={}".format(mes, len(set(i["ncm"] for i in itens)), len(itens)),
          nome)
    return itens, mes, cnpj


def extrair_xmls_de_zip(conteudo_zip: bytes) -> list[tuple[str, bytes]]:
    out: list[tuple[str, bytes]] = []
    with zipfile.ZipFile(BytesIO(conteudo_zip)) as z:
        for nome in z.namelist():
            if nome.lower().endswith(".xml"):
                out.append((os.path.basename(nome), z.read(nome)))
    log_a("INFO", "Zip", "{} XMLs extraidos".format(len(out)))
    return out


# ─────────────────────────────────────────────────────────────
#  PGDAS — leitura sem cache (arquivo pode mudar entre uploads)
# ─────────────────────────────────────────────────────────────
def ler_pgdas(file_bytes: bytes, nome: str) -> pd.DataFrame:
    """
    Le CSV/XLSX do PGDAS.
    Colunas obrigatorias: Mes, Receita_PGDAS, DAS_Pago
    Colunas opcionais:    Receita_Monofasica_PGDAS, Segregacao_PGDAS
    """
    if nome.lower().endswith(".csv"):
        df = None
        for sep, dec in ((";", ","),(","  ,".")):
            try:
                df = pd.read_csv(BytesIO(file_bytes), sep=sep, decimal=dec, dtype=str)
                if df.shape[1] >= 3:
                    break
            except Exception:
                continue
        if df is None:
            raise ValueError("Nao foi possivel ler o CSV do PGDAS.")
    else:
        try:
            df = pd.read_excel(BytesIO(file_bytes), dtype=str)
        except Exception as e:
            raise ValueError("Erro ao ler Excel do PGDAS: {}".format(e))

    # Normaliza nomes de colunas
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    low = list(df.columns)

    # Mapeamento flexivel de colunas obrigatorias
    def find_col(candidates: list[str]) -> Optional[str]:
        for c in candidates:
            if c in low: return c
        return None

    col_mes  = find_col(["mes","month","competencia","periodo"])
    col_rec  = find_col(["receita_pgdas","receita","receita_bruta","receita_total","rb"])
    col_das  = find_col(["das_pago","das","valor_das","das_recolhido"])
    col_mono = find_col(["receita_monofasica_pgdas","receita_monofasica","mono_pgdas","monofasico"])
    col_seg  = find_col(["segregacao_pgdas","segregacao","seg"])

    if not col_mes:
        raise ValueError("Coluna 'Mes' nao encontrada. Colunas: {}".format(list(df.columns)))
    if not col_rec:
        raise ValueError("Coluna 'Receita_PGDAS' nao encontrada. Colunas: {}".format(list(df.columns)))
    if not col_das:
        raise ValueError("Coluna 'DAS_Pago' nao encontrada. Colunas: {}".format(list(df.columns)))

    def norm_mes(v: str) -> str:
        v = str(v).strip()
        for fmt in ("%Y-%m", "%m/%Y", "%m-%Y", "%Y/%m", "%d/%m/%Y", "%Y%m"):
            try:
                return datetime.strptime(v, fmt).strftime("%Y-%m")
            except ValueError:
                continue
        # Tenta extrair primeiros 7 chars se formato YYYY-MM
        if len(v) >= 7 and v[4] == "-":
            return v[:7]
        return v

    def to_float(v) -> float:
        s = str(v).strip().replace("R$", "").replace(" ", "")
        if "," in s and "." in s:
            # Formato brasileiro: 1.234,56
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        try:
            return round(float(s), 2)
        except ValueError:
            return 0.0

    result = pd.DataFrame()
    result["mes"]           = df[col_mes].apply(norm_mes)
    result["receita_pgdas"] = df[col_rec].apply(to_float)
    result["das_pago"]      = df[col_das].apply(to_float)
    result["receita_monofasica_pgdas"] = df[col_mono].apply(to_float) if col_mono else 0.0
    result["segregacao_pgdas"]         = df[col_seg].apply(to_float)  if col_seg  else None

    result = result.sort_values("mes").reset_index(drop=True)
    log_a("INFO", "PGDAS",
          "{} meses: {} a {}".format(len(result), result["mes"].min(), result["mes"].max()))
    return result


# ─────────────────────────────────────────────────────────────
#  VALIDATORS
# ─────────────────────────────────────────────────────────────
def classificar_ncm(ncm: str) -> tuple[str, str]:
    """
    Classifica NCM na tabela monofasica.
    Busca por 8, 6 e 4 digitos (do mais ao menos especifico).
    """
    n = str(ncm).strip()
    if not n:
        return "INCONSISTENCIA", "NCM ausente"
    if not n.isdigit():
        return "INCONSISTENCIA", "NCM invalido: {}".format(n)
    for t in (8, 6, 4):
        chave = n[:t].ljust(8, "0")
        if chave in TABELA_NCM:
            return "MONOFASICO", TABELA_NCM[chave]
    return "NAO_MONOFASICO", "NCM fora da tabela"

def validar_tributario(ncm: str, cst_p: str, cst_c: str,
                       cfop: str, classif: str) -> dict:
    alertas: list[str] = []
    score = 0
    is_mono = classif == "MONOFASICO"
    cf = cfop.strip().replace(".", "")
    cp = cst_p.strip().lstrip("0") or "0"
    cc = cst_c.strip().lstrip("0") or "0"

    if not is_mono:
        if cp in CST_MONO or cc in CST_MONO:
            alertas.append("CST monofasico mas NCM nao consta na tabela")
            score += 40
            return {"status_tributario": "INCONSISTENTE",
                    "score_risco": min(score, 100),
                    "motivo_alerta": " | ".join(alertas)}
        return {"status_tributario": "NAO_MONOFASICO", "score_risco": 0, "motivo_alerta": ""}

    # NCM monofasico — validacao cruzada
    for label, val, raw in (("PIS", cp, cst_p), ("COFINS", cc, cst_c)):
        if val and val not in CST_MONO and val != "0":
            if val in CST_NORMAL:
                alertas.append("CST {} {} incompativel c/ monofasico".format(label, raw))
                score += 35
            else:
                alertas.append("CST {} {} nao reconhecido".format(label, raw))
                score += 15

    if cf:
        if cf in CFOP_INC:
            alertas.append("CFOP {} incompativel c/ revenda monofasica".format(cfop))
            score += 30
        elif cf not in CFOP_REV:
            alertas.append("CFOP {} nao identificado como revenda padrao".format(cfop))
            score += 10

    if cp and cc and cp != cc:
        alertas.append("CST PIS ({}) diverge do COFINS ({})".format(cst_p, cst_c))
        score += 20

    score = min(score, 100)
    status = "MONOFASICO_VALIDADO" if not alertas else "MONOFASICO_COM_RISCO"
    return {"status_tributario": status,
            "score_risco": score,
            "motivo_alerta": " | ".join(alertas) if alertas else "Validado"}


# ─────────────────────────────────────────────────────────────
#  PROCESSAMENTO DE XMLs
# ─────────────────────────────────────────────────────────────
def processar_xmls(
    arquivos: list[tuple[str, bytes]],
    aliq: float = ALIQ_ESTIMATIVA,
) -> tuple[list[dict], dict]:
    """
    Processa lista de (nome, bytes) de XMLs.
    SEM cache — bytes nao sao hashed confiavelmente pelo Streamlit.
    """
    todos: list[dict] = []
    nomes_vistos: set[str] = set()

    for nome, conteudo in arquivos:
        # Detecta duplicatas pelo nome
        if nome in nomes_vistos:
            log_a("AVISO", "Pipeline", "Arquivo duplicado ignorado", nome)
            continue
        nomes_vistos.add(nome)

        try:
            raw_itens, mes, cnpj = ler_xml_nfe(nome, conteudo)
        except ValueError as e:
            log_a("ERRO", "Pipeline", str(e), nome)
            st.warning("Erro em '{}': {}".format(nome, e))
            continue

        for i in raw_itens:
            cl, mot = classificar_ncm(i["ncm"])
            val_t   = validar_tributario(
                i["ncm"], i["cst_pis"], i["cst_cofins"], i["cfop"], cl)
            todos.append({
                **i,
                "arquivo":       nome,
                "mes":           mes,
                "cnpj":          cnpj,
                "classificacao": cl,
                "motivo_ncm":    mot,
                **val_t,
            })

    res = _resumo_geral(todos, aliq)
    return todos, res


def _resumo_geral(itens: list[dict], aliq: float = ALIQ_ESTIMATIVA) -> dict:
    tot  = sum(i["valor"] for i in itens)
    mono = sum(i["valor"] for i in itens
               if "MONOFASICO" in i.get("status_tributario", ""))
    nmon = sum(i["valor"] for i in itens
               if i.get("status_tributario") == "NAO_MONOFASICO")
    inc  = sum(i["valor"] for i in itens
               if i.get("status_tributario") == "INCONSISTENTE")
    mval = sum(i["valor"] for i in itens
               if i.get("status_tributario") == "MONOFASICO_VALIDADO")
    mris = sum(i["valor"] for i in itens
               if i.get("status_tributario") == "MONOFASICO_COM_RISCO")
    return {
        "total_geral":      r2(tot),
        "total_monofasico": r2(mono),
        "total_nao_mono":   r2(nmon),
        "total_inconsist":  r2(inc),
        "total_validado":   r2(mval),
        "total_risco":      r2(mris),
        "estimativa_recup": r2(mono * aliq),
        "pct_monofasico":   round(mono / tot, 4) if tot > 0 else 0.0,
    }


# ─────────────────────────────────────────────────────────────
#  AGRUPAMENTO MENSAL
# ─────────────────────────────────────────────────────────────
def agrupar_por_mes(itens: list[dict]) -> list[dict]:
    m: dict[str, dict] = {}
    for i in itens:
        k = i.get("mes", "SEM-DATA")
        if k not in m:
            m[k] = {"tot": 0.0, "mono": 0.0, "mono_v": 0.0, "mono_r": 0.0}
        m[k]["tot"] += i["valor"]
        st_ = i.get("status_tributario", "")
        if "MONOFASICO" in st_:
            m[k]["mono"] += i["valor"]
            if st_ == "MONOFASICO_VALIDADO":
                m[k]["mono_v"] += i["valor"]
            else:
                m[k]["mono_r"] += i["valor"]

    out: list[dict] = []
    for k in sorted(x for x in m if x != "SEM-DATA") + \
             [x for x in m if x == "SEM-DATA"]:
        t  = round(m[k]["tot"],  2)
        mo = round(m[k]["mono"], 2)
        out.append({
            "mes":                   k,
            "receita_total":         t,
            "receita_monofasica":    mo,
            "receita_mono_validada": round(m[k]["mono_v"], 2),
            "receita_mono_risco":    round(m[k]["mono_r"], 2),
            "receita_tributavel":    round(t - mo, 2),
        })
    return out


# ─────────────────────────────────────────────────────────────
#  APURACAO PGDAS-D
# ─────────────────────────────────────────────────────────────
def _faixa(rbt12: float, tab: list[dict]) -> Optional[dict]:
    if rbt12 <= 0: return tab[0]
    for f in tab:
        if rbt12 <= f["limite"]: return f
    return None

def _aliq_ef(rbt12: float, faixa: dict) -> float:
    if rbt12 <= 0: return 0.0
    return (rbt12 * faixa["aliquota"] - faixa["deducao"]) / rbt12

def _pct_pc(fn: int, anexo: str) -> float:
    r = REPARTICAO.get(anexo, {}).get(fn, {})
    return r.get("pis", 0.0) + r.get("cofins", 0.0)

def _status_credito(cb: Optional[float], cr: Optional[float]) -> dict:
    if cb is None or cr is None:
        return {"status_credito": "NAO_RECUPERAVEL",
                "credito_final": None,
                "motivo_credito": "Calculo indisponivel"}
    if cb < 0:
        log_a("RISCO", "Credito", "Credito bruto negativo", brl(cb))
        return {"status_credito": "RISCO_FISCAL",
                "credito_final": 0.0,
                "motivo_credito": "DAS Correto > DAS Usado"}
    if cr <= 0:
        return {"status_credito": "NAO_RECUPERAVEL",
                "credito_final": 0.0,
                "motivo_credito": "Credito real zerado"}
    return {"status_credito": "RECUPERAVEL",
            "credito_final": round(cr, 2),
            "motivo_credito": ""}

def _status_segregacao(mono_xml: float,
                       mono_pgdas: Optional[float]) -> str:
    if mono_pgdas is None:
        return "SEM_DADO_PGDAS"
    if mono_xml <= 0:
        return "SEM_SEGREGACAO"
    ratio = mono_pgdas / mono_xml
    if ratio < SEG_PARCIAL_MIN:  return "SEM_SEGREGACAO"
    if ratio < SEG_TOTAL_MIN:    return "SEGREGACAO_PARCIAL"
    return "SEGREGACAO_TOTAL"

def calcular_rbt12(agrup: list[dict], rbt12_ini: float = 0.0,
                   df_pgdas: Optional[pd.DataFrame] = None) -> dict[str, float]:
    """
    RBT12 rolling de 12 meses.
    Usa receita_pgdas como fonte primaria (mais preciso que XML).
    """
    # Indice mes → receita_pgdas
    pgdas_r: dict[str, float] = {}
    if df_pgdas is not None and not df_pgdas.empty:
        for _, row in df_pgdas.iterrows():
            pgdas_r[str(row["mes"])] = safe_float(row["receita_pgdas"])

    res: dict[str, float] = {}
    hist: list[tuple[str, float]] = []
    saldo = rbt12_ini

    for row in agrup:
        mes = row["mes"]
        res[mes] = round(saldo, 2)
        rec = pgdas_r.get(mes, row["receita_total"])
        hist.append((mes, rec))
        saldo += rec
        if len(hist) > 12:
            _, antiga = hist.pop(0)
            saldo -= antiga

    return res

def apurar_periodo(
    agrup: list[dict],
    nome_anexo: str,
    rbt12_ini: float = 0.0,
    df_pgdas: Optional[pd.DataFrame] = None,
) -> list[dict]:
    """
    Pipeline de apuracao mensal. SEM cache (df_pgdas nao e serializavel).
    Combina dados dos XMLs com dados reais do PGDAS quando disponiveis.
    """
    tabela   = TABELAS_SIMPLES[nome_anexo]
    rbt12map = calcular_rbt12(agrup, rbt12_ini, df_pgdas)

    # Indice PGDAS por mes — acesso direto sem risco de conversao
    pgdas_idx: dict[str, dict] = {}
    if df_pgdas is not None and not df_pgdas.empty:
        for _, row in df_pgdas.iterrows():
            mes_k = str(row["mes"]).strip()
            pgdas_idx[mes_k] = {
                "receita_pgdas":             safe_float(row.get("receita_pgdas")),
                "das_pago":                  safe_float(row.get("das_pago")),
                "receita_monofasica_pgdas":  safe_float(row.get("receita_monofasica_pgdas")),
                "segregacao_pgdas":          safe_float(row.get("segregacao_pgdas")),
            }

    out: list[dict] = []
    for row in agrup:
        mes     = row["mes"]
        rbt12   = rbt12map.get(mes, 0.0)
        faixa   = _faixa(rbt12, tabela)
        pr      = pgdas_idx.get(mes)
        tem_pgdas = pr is not None

        # Campos do PGDAS (None quando ausente)
        rec_pgdas  = pr["receita_pgdas"]               if pr else None
        das_real   = pr["das_pago"]                    if pr else None
        mono_pgdas = pr["receita_monofasica_pgdas"]    if pr else None
        seg_pgdas  = pr["segregacao_pgdas"]             if pr else None

        # Divergencia XML vs PGDAS
        div = None
        alerta_div = False
        if rec_pgdas is not None and rec_pgdas > 0:
            div = round(abs(row["receita_total"] - rec_pgdas) / rec_pgdas, 4)
            alerta_div = div > DIVERGENCIA_MAX
            if alerta_div:
                log_a("AVISO", "Apuracao",
                      "Divergencia {:.1f}% em {}".format(div * 100, mes),
                      "XML={} PGDAS={}".format(brl(row["receita_total"]), brl(rec_pgdas)))

        # Status segregacao
        mono_pgdas_v = mono_pgdas if (mono_pgdas is not None and mono_pgdas > 0) else None
        status_seg = _status_segregacao(row["receita_monofasica"], mono_pgdas_v)

        if faixa is None:
            out.append({
                **row,
                "rbt12": rbt12, "faixa": "ACIMA",
                "aliquota_nominal": None, "aliquota_efetiva": None,
                "receita_pgdas": rec_pgdas, "das_pago_real": das_real,
                "receita_monofasica_pgdas": mono_pgdas_v, "segregacao_pgdas": seg_pgdas,
                "divergencia_pct": div, "alerta_divergencia": alerta_div,
                "tem_pgdas": tem_pgdas, "status_segregacao": status_seg,
                "das_estimado": None, "das_usado": None, "fonte_das": "–",
                "das_correto": None, "credito_bruto": None,
                "pct_pis_cofins_val": None, "credito_real": None,
                "status_credito": "NAO_RECUPERAVEL", "credito_final": 0.0,
                "motivo_credito": "RBT12 acima de R$ 4,8M", "alerta": "Fora do Simples",
            })
            log_a("RISCO", "Apuracao", "RBT12 fora do Simples", "Mes={}".format(mes))
            continue

        ae      = _aliq_ef(rbt12, faixa)
        das_est = r2(row["receita_total"]      * ae)
        das_cor = r2(row["receita_tributavel"] * ae)

        # Fonte do DAS: real quando disponivel, estimado como fallback
        if tem_pgdas and das_real is not None and das_real > 0:
            das_usado = r2(das_real)
            fonte = "REAL"
        else:
            das_usado = das_est
            fonte = "ESTIMADO"

        cb  = r2((das_usado - das_cor) if das_usado is not None else None)
        ppc = _pct_pc(faixa["faixa"], nome_anexo)
        cr  = r2(cb * ppc) if cb is not None else None
        sc  = _status_credito(cb, cr)

        out.append({
            **row,
            "rbt12":                     rbt12,
            "faixa":                     faixa["faixa"],
            "aliquota_nominal":          faixa["aliquota"],
            "aliquota_efetiva":          ae,
            "receita_pgdas":             rec_pgdas,
            "das_pago_real":             das_real,
            "receita_monofasica_pgdas":  mono_pgdas_v,
            "segregacao_pgdas":          seg_pgdas,
            "divergencia_pct":           div,
            "alerta_divergencia":        alerta_div,
            "tem_pgdas":                 tem_pgdas,
            "status_segregacao":         status_seg,
            "das_estimado":              das_est,
            "das_usado":                 das_usado,
            "fonte_das":                 fonte,
            "das_correto":               das_cor,
            "credito_bruto":             cb,
            "pct_pis_cofins_val":        ppc,
            "credito_real":              cr,
            **sc,
            "alerta":                    "",
        })

    return out


# ─────────────────────────────────────────────────────────────
#  SCORE E RESUMO EXECUTIVO
# ─────────────────────────────────────────────────────────────
def score_oportunidade(res: dict, apuracao: list[dict]) -> dict:
    s = 0
    pctm = res.get("pct_monofasico", 0.0)
    if   pctm >= 0.60: s += 40
    elif pctm >= 0.30: s += 25
    elif pctm >= 0.10: s += 10
    cred = sum(r.get("credito_final", 0) or 0 for r in apuracao)
    if   cred >= 10_000: s += 30
    elif cred >= 3_000:  s += 20
    elif cred >= 500:    s += 10
    n_risco = sum(1 for r in apuracao if r.get("status_credito") == "RISCO_FISCAL")
    if len(apuracao) > 0:
        taxa = n_risco / len(apuracao)
        if taxa == 0:    s += 20
        elif taxa < 0.2: s += 10
    ndiv = sum(1 for r in apuracao if r.get("alerta_divergencia"))
    if   ndiv == 0: s += 10
    elif ndiv <= 2: s += 5
    nivel = "ALTA" if s >= 70 else ("MEDIA" if s >= 40 else "BAIXA")
    emoji = {"ALTA": "🟢", "MEDIA": "🟡", "BAIXA": "🔴"}[nivel]
    return {"score": s, "nivel": nivel, "emoji": emoji,
            "credito_total": round(cred, 2)}

def texto_exec(res: dict, apuracao: list[dict], score: dict, n_xmls: int) -> str:
    n_seg = sum(1 for r in apuracao if r.get("status_segregacao") == "SEGREGACAO_TOTAL")
    n_par = sum(1 for r in apuracao if r.get("status_segregacao") == "SEGREGACAO_PARCIAL")
    n_sem = sum(1 for r in apuracao if r.get("status_segregacao") == "SEM_SEGREGACAO")
    n_div = sum(1 for r in apuracao if r.get("alerta_divergencia"))
    meses_cred = [r for r in apuracao if (r.get("credito_final") or 0) > 0]
    return "\n".join([
        "Foram analisados {} arquivo(s) XML, {} mes(es).".format(n_xmls, len(apuracao)),
        "",
        "OPORTUNIDADE:",
        "  Receita monofasica: {} ({:.1f}% do total)".format(
            brl(res["total_monofasico"]), res["pct_monofasico"] * 100),
        "  Credito potencial PIS/COFINS: {}".format(brl(score["credito_total"])),
        "  Meses com credito recuperavel: {}".format(len(meses_cred)),
        "",
        "SEGREGACAO:",
        "  Total: {} | Parcial: {} | Sem segregacao: {}".format(n_seg, n_par, n_sem),
        "",
        "ALERTAS:",
        "  Divergencias XML vs PGDAS (>5%): {} mes(es)".format(n_div),
        "",
        "SCORE: {} {} ({}/100)".format(score["nivel"], score["emoji"], score["score"]),
        "",
        "Base legal: LC 123/2006 | CGSN 140/2018 | Lei 10.147/2000",
        "Aviso: Estimativa preliminar. Validar com contador habilitado.",
    ])


# ─────────────────────────────────────────────────────────────
#  AUDITORIA
# ─────────────────────────────────────────────────────────────
def motor_auditoria(itens: list[dict], apuracao: list[dict],
                    df_pgdas: Optional[pd.DataFrame]) -> list[dict]:
    alertas: list[dict] = []
    def add(sev, mes, desc, impacto="", cat="Geral"):
        alertas.append({"severidade": sev, "mes": mes, "categoria": cat,
                        "descricao": desc, "impacto": impacto})
        log_a("AVISO" if sev in ("CRITICO","ALTO") else "INFO",
              "Auditoria", "[{}] {}".format(sev, desc), mes)

    for r in apuracao:
        if r.get("alerta_divergencia"):
            add("ALTO", fmt_mes(r["mes"]),
                "Divergencia {:.1f}% entre XML e PGDAS".format(r["divergencia_pct"] * 100),
                "XML={} | PGDAS={}".format(brl(r.get("receita_total")), brl(r.get("receita_pgdas"))),
                "Divergencia")
        if r.get("status_credito") == "RISCO_FISCAL":
            add("CRITICO", fmt_mes(r["mes"]),
                "Credito bruto negativo – risco fiscal",
                "DAS Usado={} | DAS Correto={}".format(
                    brl(r.get("das_usado")), brl(r.get("das_correto"))), "Credito")
        if r.get("status_segregacao") == "SEM_SEGREGACAO" and r.get("receita_monofasica", 0) > 0:
            add("ALTO", fmt_mes(r["mes"]),
                "Empresa nao segregou receita monofasica no PGDAS",
                "Monofasico XML={}".format(brl(r.get("receita_monofasica"))), "Segregacao")
        elif r.get("status_segregacao") == "SEGREGACAO_PARCIAL":
            add("MEDIO", fmt_mes(r["mes"]), "Segregacao parcial detectada",
                "XML={} | PGDAS={}".format(
                    brl(r.get("receita_monofasica")), brl(r.get("receita_monofasica_pgdas"))),
                "Segregacao")

    if df_pgdas is not None and not df_pgdas.empty:
        meses_pgdas = set(df_pgdas["mes"].astype(str).tolist())
        for r in apuracao:
            if r["mes"] != "SEM-DATA" and r["mes"] not in meses_pgdas:
                add("MEDIO", fmt_mes(r["mes"]),
                    "Mes nos XMLs mas ausente no PGDAS",
                    "Rec. XML={}".format(brl(r.get("receita_total"))), "PGDAS")

    alto_risco = [i for i in itens if i.get("score_risco", 0) >= 50]
    if alto_risco:
        add("ALTO", "Varios itens",
            "{} item(ns) com score tributario >= 50".format(len(alto_risco)),
            "Valor: {}".format(brl(sum(i["valor"] for i in alto_risco))),
            "Risco Tributario")

    sem_data = [i for i in itens if i.get("mes") == "SEM-DATA"]
    if sem_data:
        add("MEDIO", "SEM-DATA",
            "{} item(ns) sem data de emissao".format(len(sem_data)),
            "Valor: {}".format(brl(sum(i["valor"] for i in sem_data))), "XML")

    sev_ord = {"CRITICO": 0, "ALTO": 1, "MEDIO": 2, "BAIXO": 3}
    return sorted(alertas, key=lambda x: sev_ord.get(x["severidade"], 9))


# ─────────────────────────────────────────────────────────────
#  DASHBOARD — graficos nativos + plotly opcional
# ─────────────────────────────────────────────────────────────
def graficos(res: dict, apuracao: list[dict], itens: list[dict]) -> None:
    t1, t2, t3, t4 = st.tabs(
        ["Composicao", "Evolucao Mensal", "Credito por Mes", "Ranking NCM"])

    valids = [r for r in apuracao if r.get("receita_total") is not None]

    with t1:
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Composicao do Faturamento**")
            df_comp = pd.DataFrame({
                "Categoria": ["Validado","Com Risco","Nao Monofasico","Inconsistente"],
                "Valor":     [res["total_validado"], res["total_risco"],
                              res["total_nao_mono"],  res["total_inconsist"]],
            }).set_index("Categoria")
            if PLOTLY_OK:
                fig = px.pie(df_comp.reset_index(), names="Categoria", values="Valor",
                             color_discrete_sequence=["#1a7a4a","#b8860b","#2d5a8e","#8b1a1a"],
                             hole=0.4)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.bar_chart(df_comp)

        with c2:
            st.markdown("**Status de Segregacao (meses)**")
            seg_vals = {
                "Seg. Total":   sum(1 for r in apuracao if r.get("status_segregacao") == "SEGREGACAO_TOTAL"),
                "Seg. Parcial": sum(1 for r in apuracao if r.get("status_segregacao") == "SEGREGACAO_PARCIAL"),
                "Sem Seg.":     sum(1 for r in apuracao if r.get("status_segregacao") == "SEM_SEGREGACAO"),
                "Sem Dado":     sum(1 for r in apuracao if r.get("status_segregacao") == "SEM_DADO_PGDAS"),
            }
            if PLOTLY_OK:
                fig2 = px.pie(names=list(seg_vals.keys()), values=list(seg_vals.values()),
                              color_discrete_sequence=["#1a7a4a","#b8860b","#8b1a1a","#6c757d"],
                              hole=0.4)
                st.plotly_chart(fig2, use_container_width=True)
            else:
                st.bar_chart(pd.DataFrame.from_dict(seg_vals, orient="index", columns=["Meses"]))

    with t2:
        if valids:
            df_ev = pd.DataFrame({
                "Mes":    [fmt_mes(r["mes"])          for r in valids],
                "XML":    [r["receita_total"]          for r in valids],
                "PGDAS":  [r.get("receita_pgdas") or 0 for r in valids],
                "Monof.": [r["receita_monofasica"]     for r in valids],
                "Trib.":  [r["receita_tributavel"]     for r in valids],
            }).set_index("Mes")
            if PLOTLY_OK:
                fig = px.line(df_ev.reset_index(), x="Mes",
                              y=["XML","PGDAS","Monof.","Trib."],
                              title="Evolucao Mensal da Receita", markers=True)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.line_chart(df_ev)

    with t3:
        cred_valids = [r for r in apuracao if r.get("credito_final") is not None]
        if cred_valids:
            df_cred = pd.DataFrame({
                "Mes":    [fmt_mes(r["mes"])               for r in cred_valids],
                "Credito":[r.get("credito_final", 0) or 0  for r in cred_valids],
            }).set_index("Mes")
            if PLOTLY_OK:
                cor_mapa = {"RECUPERAVEL":"#1a7a4a","NAO_RECUPERAVEL":"#2d5a8e","RISCO_FISCAL":"#8b1a1a"}
                fig = px.bar(
                    x=[fmt_mes(r["mes"]) for r in cred_valids],
                    y=[r.get("credito_final", 0) or 0 for r in cred_valids],
                    color=[r.get("status_credito","") for r in cred_valids],
                    color_discrete_map=cor_mapa,
                    title="Credito Final por Mes",
                    labels={"x":"Mes","y":"R$","color":"Status"},
                )
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.bar_chart(df_cred)

    with t4:
        df_i = pd.DataFrame(itens)
        if "classificacao" in df_i.columns:
            df_ncm = (df_i[df_i["classificacao"] == "MONOFASICO"]
                      .groupby("ncm")["valor"].sum()
                      .reset_index()
                      .sort_values("valor", ascending=False)
                      .head(15))
            if not df_ncm.empty:
                if PLOTLY_OK:
                    fig = px.bar(df_ncm, x="valor", y="ncm", orientation="h",
                                 title="Top 15 NCMs Monofasicos", labels={"valor":"R$","ncm":"NCM"})
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.bar_chart(df_ncm.set_index("ncm"))


# ─────────────────────────────────────────────────────────────
#  EXPORTS
# ─────────────────────────────────────────────────────────────
def gerar_csv(apuracao: list[dict]) -> bytes:
    rows = []
    for r in apuracao:
        rows.append({
            "Mes":            r.get("mes",""),
            "RBT12":          r.get("rbt12",""),
            "Faixa":          r.get("faixa",""),
            "Aliq. Efetiva":  "{:.2f}%".format(r["aliquota_efetiva"]*100).replace(".",",")
                              if r.get("aliquota_efetiva") is not None else "",
            "Rec. XML":       r.get("receita_total",""),
            "Rec. PGDAS":     r.get("receita_pgdas","") or "",
            "Divergencia":    "{:.2f}%".format(r["divergencia_pct"]*100).replace(".",",")
                              if r.get("divergencia_pct") is not None else "",
            "Rec. Monof.":    r.get("receita_monofasica",""),
            "Rec. Monof. PGDAS": r.get("receita_monofasica_pgdas","") or "",
            "Segregacao":     r.get("status_segregacao",""),
            "Rec. Tributavel":r.get("receita_tributavel",""),
            "DAS Usado":      r.get("das_usado",""),
            "Fonte DAS":      r.get("fonte_das",""),
            "DAS Correto":    r.get("das_correto",""),
            "Cred. Bruto":    r.get("credito_bruto",""),
            "% PIS+COF":      "{:.2f}%".format(r["pct_pis_cofins_val"]*100).replace(".",",")
                              if r.get("pct_pis_cofins_val") is not None else "",
            "Cred. Real":     r.get("credito_real",""),
            "Cred. Final":    r.get("credito_final",""),
            "Status Credito": r.get("status_credito",""),
        })
    return pd.DataFrame(rows).to_csv(index=False, sep=";", decimal=",").encode("utf-8")

def gerar_excel(itens: list[dict], res: dict, apuracao: list[dict],
                score: dict, alertas_a: list[dict], logs: list[dict]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as w:
        # Aba 1 – Resumo
        pd.DataFrame([
            {"Indicador":"Faturamento Total",              "Valor":brl(res["total_geral"])},
            {"Indicador":"Faturamento Monofasico",         "Valor":brl(res["total_monofasico"])},
            {"Indicador":"  Validado",                     "Valor":brl(res["total_validado"])},
            {"Indicador":"  Com risco tributario",         "Valor":brl(res["total_risco"])},
            {"Indicador":"Faturamento Nao Monofasico",     "Valor":brl(res["total_nao_mono"])},
            {"Indicador":"% Monofasico",                   "Valor":pct_s(res["pct_monofasico"])},
            {"Indicador":"Credito Potencial PIS/COFINS",   "Valor":brl(score["credito_total"])},
            {"Indicador":"Score de Oportunidade",          "Valor":"{} ({}/100)".format(score["nivel"],score["score"])},
            {"Indicador":"Total de Alertas de Auditoria",  "Valor":str(len(alertas_a))},
        ]).to_excel(w, sheet_name="1.Resumo Executivo", index=False)

        # Aba 2 – Apuracao
        pd.DataFrame([{
            "Mes":fmt_mes(r["mes"]),
            "RBT12":r.get("rbt12"),
            "Faixa":r.get("faixa","–"),
            "Aliq. Nominal":pct_s(r.get("aliquota_nominal")),
            "Aliq. Efetiva":pct_s(r.get("aliquota_efetiva"),4),
            "Rec. XML":r.get("receita_total"),
            "Rec. PGDAS":r.get("receita_pgdas"),
            "Divergencia":pct_s(r.get("divergencia_pct")) if r.get("divergencia_pct") is not None else "–",
            "Rec. Monof.":r.get("receita_monofasica"),
            "Rec. Monof. PGDAS":r.get("receita_monofasica_pgdas"),
            "Segregacao":r.get("status_segregacao","–"),
            "Rec. Tributavel":r.get("receita_tributavel"),
            "DAS Usado":r.get("das_usado"),
            "Fonte DAS":r.get("fonte_das","–"),
            "DAS Correto":r.get("das_correto"),
            "% PIS+COF":pct_s(r.get("pct_pis_cofins_val")),
            "Cred. Bruto":r.get("credito_bruto"),
            "Cred. Final":r.get("credito_final"),
            "Status Credito":r.get("status_credito","–"),
        } for r in apuracao]).to_excel(w, sheet_name="2.Apuracao Mensal", index=False)

        # Aba 3 – Itens
        pd.DataFrame([{
            "Mes":i.get("mes",""),"Arquivo":i.get("arquivo",""),
            "Descricao":i.get("descricao",""),"NCM":i.get("ncm",""),
            "CFOP":i.get("cfop",""),"CST PIS":i.get("cst_pis",""),
            "CST COF":i.get("cst_cofins",""),"Valor":i.get("valor"),
            "Class. NCM":i.get("classificacao",""),
            "Status Trib.":i.get("status_tributario",""),
            "Score":i.get("score_risco",""),"Motivo":i.get("motivo_alerta",""),
        } for i in itens]).to_excel(w, sheet_name="3.Itens Classificados", index=False)

        # Aba 4 – Divergencias
        divs = [r for r in apuracao if r.get("divergencia_pct") is not None]
        (pd.DataFrame([{
            "Mes":fmt_mes(r["mes"]),"Rec. XML":brl(r.get("receita_total")),
            "Rec. PGDAS":brl(r.get("receita_pgdas")),
            "Divergencia":pct_s(r.get("divergencia_pct")),
            "Alerta":("SIM" if r.get("alerta_divergencia") else "NAO"),
            "Segregacao":r.get("status_segregacao","–"),
        } for r in divs]) if divs else pd.DataFrame([{"Info":"Sem divergencias"}])
        ).to_excel(w, sheet_name="4.Divergencias PGDAS", index=False)

        # Aba 5 – Auditoria
        (pd.DataFrame(alertas_a) if alertas_a
         else pd.DataFrame([{"Info":"Sem alertas"}])
        ).to_excel(w, sheet_name="5.Auditoria Tributaria", index=False)

        # Aba 6 – Inconsistencias
        inc = [i for i in itens if i.get("status_tributario") in
               ("INCONSISTENTE","MONOFASICO_COM_RISCO")]
        (pd.DataFrame([{
            "Mes":i.get("mes",""),"Descricao":i.get("descricao",""),
            "NCM":i.get("ncm",""),"CFOP":i.get("cfop",""),
            "CST PIS":i.get("cst_pis",""),"CST COF":i.get("cst_cofins",""),
            "Valor":brl(i.get("valor")),"Status":i.get("status_tributario",""),
            "Score":i.get("score_risco",""),"Motivo":i.get("motivo_alerta",""),
        } for i in inc]) if inc else pd.DataFrame([{"Info":"Sem inconsistencias"}])
        ).to_excel(w, sheet_name="6.Inconsistencias NCM", index=False)

    # Estilizacao basica com openpyxl
    if OPENPYXL_OK:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = openpyxl.load_workbook(BytesIO(output.getvalue()))
        hdr = PatternFill("solid", fgColor="1E3A5F")
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font      = Font(bold=True, color="FFFFFF", size=10)
                cell.fill      = hdr
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = ws["A2"]
            for col in ws.columns:
                ml = max(len(str(c.value or "")) for c in col) + 3
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml, 50)
        buf2 = BytesIO()
        wb.save(buf2)
        return buf2.getvalue()

    return output.getvalue()

def gerar_pdf(res: dict, apuracao: list[dict], score: dict, txt_exec: str) -> bytes:
    if not REPORTLAB_OK:
        return b""
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    azul   = rl_colors.HexColor("#1E3A5F")

    def H1(t): return Paragraph(t, ParagraphStyle("h1", parent=styles["Heading1"],
                                 textColor=azul, fontSize=16))
    def H2(t): return Paragraph(t, ParagraphStyle("h2", parent=styles["Heading2"],
                                 textColor=azul, fontSize=12))
    def P(t):  return Paragraph(t, styles["Normal"])

    dados = [["Indicador","Valor"],
             ["Faturamento Total", brl(res["total_geral"])],
             ["Receita Monofasica", brl(res["total_monofasico"])],
             ["% Monofasico", pct_s(res["pct_monofasico"])],
             ["Credito Potencial", brl(score["credito_total"])],
             ["Score", "{} ({}/100)".format(score["nivel"], score["score"])]]
    t = Table(dados, colWidths=[10*cm, 6*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0), azul),
        ("TEXTCOLOR",(0,0),(-1,0), rl_colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.white, rl_colors.HexColor("#EBF0F7")]),
        ("GRID",(0,0),(-1,-1),0.5, rl_colors.HexColor("#CCCCCC")),
        ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
    ]))
    story = [H1("PIS/COFINS Pro – Relatorio Executivo"),
             P("Emitido: {}  |  v{}".format(datetime.now().strftime("%d/%m/%Y %H:%M"), VERSAO)),
             HRFlowable(width="100%", thickness=2, color=azul), Spacer(1,0.4*cm),
             H2("Indicadores Principais"), t, Spacer(1,0.4*cm),
             H2("Resumo Executivo")]
    for linha in txt_exec.split("\n"):
        if linha.strip():
            story.append(P(linha))
    story += [Spacer(1,0.4*cm),
              HRFlowable(width="100%",thickness=1,color=rl_colors.HexColor("#CCCCCC")),
              P("<i>Aviso: Analise preliminar. LC 123/2006.</i>")]
    doc.build(story)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
#  INTERFACE STREAMLIT
# ─────────────────────────────────────────────────────────────
def main() -> None:
    st.set_page_config(page_title="PIS/COFINS Pro", page_icon="📊",
                       layout="wide", initial_sidebar_state="expanded")
    st.markdown("""<style>
    .saas-hdr{background:linear-gradient(135deg,#1E3A5F,#2d5a8e);color:white;
              padding:18px 22px;border-radius:12px;margin-bottom:14px;}
    .saas-hdr h1{margin:0;font-size:24px;font-weight:800;}
    .saas-hdr p{margin:3px 0 0;opacity:.8;font-size:13px;}
    .sec{font-size:16px;font-weight:700;color:#1E3A5F;
         border-left:4px solid #2d5a8e;padding-left:10px;margin:16px 0 8px;}
    .score-box{border-radius:14px;padding:20px;text-align:center;color:white;}
    div[data-testid="stMetricValue"]{font-size:19px!important;font-weight:700!important;}
    </style>""", unsafe_allow_html=True)

    st.markdown("""<div class="saas-hdr">
      <h1>📊 PIS/COFINS Pro &nbsp;|&nbsp; Recuperacao Tributaria – Simples Nacional</h1>
      <p>NCM · CST · CFOP · PGDAS-D · RBT12 Dinamico · v{} {}</p>
    </div>""".format(VERSAO, "| Plotly ✅" if PLOTLY_OK else "| Graficos nativos"), unsafe_allow_html=True)

    clear_logs()

    # ── SIDEBAR ───────────────────────────────────────────────
    with st.sidebar:
        st.markdown("### ⚙️ Configuracoes")
        nome_anexo = st.selectbox("Anexo do Simples", list(TABELAS_SIMPLES.keys()))
        rbt12_ini  = st.number_input("RBT12 anterior ao periodo (R$)",
                                     min_value=0.0, value=360_000.0,
                                     step=1_000.0, format="%.2f",
                                     help="Receita dos 12 meses antes do primeiro XML.")
        usar_pgdas = st.checkbox("Usar dados reais do PGDAS", value=False)
        aliq_pct   = st.number_input("Aliquota PIS+COFINS estimativa (%)",
                                     min_value=0.0, max_value=100.0,
                                     value=9.25, step=0.05, format="%.2f")
        aliq_dec   = aliq_pct / 100
        st.markdown("---")
        st.caption("`{}` NCMs monofasicos na tabela".format(len(TABELA_NCM)))
        if not PLOTLY_OK:
            st.info("Graficos nativos ativos.\n`pip install plotly` para graficos avancados.")
        if not REPORTLAB_OK:
            st.info("PDF desativado.\n`pip install reportlab`")

    # ── UPLOAD XMLs ───────────────────────────────────────────
    st.markdown('<div class="sec">Upload de Documentos Fiscais</div>', unsafe_allow_html=True)
    col_x, col_p = st.columns(2)

    with col_x:
        st.markdown("**XMLs de NF-e ou arquivo .zip**")
        up_xmls = st.file_uploader(
            "Selecione XMLs ou .zip", type=["xml","zip"],
            accept_multiple_files=True, key="xmls",
            help="Limite: {}MB por arquivo".format(UPLOAD_MAX_MB))

    df_pgdas = None
    with col_p:
        if usar_pgdas:
            st.markdown("**Planilha PGDAS/DAS**")
            with st.expander("Formato esperado"):
                st.markdown("""
Colunas obrigatorias (nomes flexiveis):

| Mes | Receita_PGDAS | DAS_Pago |
|---|---|---|
| 2025-01 | 135000,00 | 5940,00 |

Colunas opcionais: `Receita_Monofasica_PGDAS`, `Segregacao_PGDAS`
                """)
                st.download_button("Baixar template CSV",
                    data="Mes;Receita_PGDAS;DAS_Pago;Receita_Monofasica_PGDAS\n2025-01;135000,00;5940,00;81000,00\n".encode(),
                    file_name="template_pgdas_v8.csv", mime="text/csv")
            up_pgdas = st.file_uploader("CSV ou Excel do PGDAS",
                                        type=["csv","xlsx","xls"], key="pgdas_up")
            if up_pgdas:
                try:
                    df_pgdas = ler_pgdas(up_pgdas.read(), up_pgdas.name)
                    st.success("PGDAS carregado: {} mes(es) | {} a {}".format(
                        len(df_pgdas), fmt_mes(df_pgdas["mes"].min()),
                        fmt_mes(df_pgdas["mes"].max())))
                    with st.expander("Visualizar PGDAS carregado"):
                        st.dataframe(df_pgdas, hide_index=True, use_container_width=True)
                except ValueError as e:
                    st.error(str(e))
        else:
            st.info("Ative 'Usar dados reais do PGDAS' na sidebar para integrar o PGDAS.")

    if not up_xmls:
        st.info("📂 Aguardando upload dos XMLs para iniciar a analise.")
        st.stop()

    # ── COLETA DE ARQUIVOS ────────────────────────────────────
    arquivos: list[tuple[str, bytes]] = []
    for f in up_xmls:
        conteudo = f.read()
        if f.name.lower().endswith(".zip"):
            try:
                arquivos.extend(extrair_xmls_de_zip(conteudo))
            except Exception as e:
                st.warning("Erro no zip '{}': {}".format(f.name, e))
        else:
            arquivos.append((f.name, conteudo))

    if not arquivos:
        st.error("Nenhum XML encontrado nos arquivos enviados.")
        st.stop()

    # ── PROCESSAMENTO ─────────────────────────────────────────
    prog = st.progress(0, "Lendo e classificando {} arquivo(s)...".format(len(arquivos)))
    itens, res = processar_xmls(arquivos, aliq_dec)
    prog.progress(40, "Agrupando por mes...")

    if not itens:
        st.error("Nenhum item extraido dos XMLs. Verifique os arquivos.")
        st.stop()

    agrup    = agrupar_por_mes(itens)
    prog.progress(60, "Apurando PGDAS-D...")
    apuracao = apurar_periodo(agrup, nome_anexo, rbt12_ini, df_pgdas)
    prog.progress(80, "Gerando analises...")
    sc       = score_oportunidade(res, apuracao)
    alertas  = motor_auditoria(itens, apuracao, df_pgdas)
    txt      = texto_exec(res, apuracao, sc, len(arquivos))
    df_itens = pd.DataFrame(itens)
    prog.progress(100, "Concluido!")
    prog.empty()

    # ── DIAGNOSTICO DE LEITURA ────────────────────────────────
    meses_encontrados = sorted(set(i["mes"] for i in itens))
    with st.expander("Diagnostico de leitura ({} itens | {} mes(es))".format(
            len(itens), len(meses_encontrados))):
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            st.markdown("**Meses identificados nos XMLs:**")
            for m in meses_encontrados:
                cnt   = sum(1 for i in itens if i["mes"] == m)
                valor = sum(i["valor"] for i in itens if i["mes"] == m)
                st.write("- **{}** → {} itens | {}".format(fmt_mes(m), cnt, brl(valor)))
        with col_d2:
            st.markdown("**Classificacao NCM:**")
            for st_, cnt in pd.DataFrame(itens)["status_tributario"].value_counts().items():
                st.write("- {}: {}".format(st_, cnt))

    # Nivel de analise
    nivel = "2 – Apuracao Real (PGDAS+XML)" if (usar_pgdas and df_pgdas is not None) else "1 – Triagem (XML apenas)"
    st.info("🔎 Nivel de Analise: **{}**".format(nivel))

    # ── SCORE ─────────────────────────────────────────────────
    st.markdown('<div class="sec">Score de Oportunidade</div>', unsafe_allow_html=True)
    _, sc_col, _ = st.columns([2,3,2])
    with sc_col:
        cor_map = {"ALTA":"#1a7a4a","MEDIA":"#b8860b","BAIXA":"#8b1a1a"}
        st.markdown("""<div class="score-box" style="background:{};">
            <div style="font-size:44px">{}</div>
            <div style="font-size:28px;font-weight:800">{} OPORTUNIDADE</div>
            <div style="font-size:17px;opacity:.9">Score: {}/100</div>
            <div style="font-size:15px;margin-top:6px">Credito Potencial: {}</div>
        </div>""".format(cor_map[sc["nivel"]], sc["emoji"],
                         sc["nivel"], sc["score"], brl(sc["credito_total"])),
                    unsafe_allow_html=True)

    # ── METRICAS ──────────────────────────────────────────────
    st.markdown('<div class="sec">Indicadores Executivos</div>', unsafe_allow_html=True)
    m1,m2,m3,m4,m5,m6 = st.columns(6)
    m1.metric("Total Analisado",   brl(res["total_geral"]))
    m2.metric("Monofasico",        brl(res["total_monofasico"]),
              pct_s(res["pct_monofasico"]))
    m3.metric("Validado",          brl(res["total_validado"]))
    m4.metric("Com Risco Trib.",   brl(res["total_risco"]))
    m5.metric("Credito Potencial", brl(sc["credito_total"]))
    m6.metric("Alertas",           str(len(alertas)))

    # ── AUDITORIA ─────────────────────────────────────────────
    if alertas:
        st.markdown('<div class="sec">Auditoria Tributaria</div>', unsafe_allow_html=True)
        icones = {"CRITICO":"🔴","ALTO":"🟠","MEDIO":"🟡","BAIXO":"🔵"}
        for al in alertas:
            ic  = icones.get(al["severidade"],"⚪")
            txt_a = "{} **[{}]** {} – {} | {}".format(
                ic, al["severidade"], al["mes"],
                al["descricao"], al["impacto"])
            if al["severidade"] in ("CRITICO","ALTO"): st.error(txt_a)
            elif al["severidade"] == "MEDIO": st.warning(txt_a)
            else: st.info(txt_a)

    # ── DASHBOARD ─────────────────────────────────────────────
    st.markdown('<div class="sec">Dashboard Executivo</div>', unsafe_allow_html=True)
    graficos(res, apuracao, itens)

    # ── APURACAO MENSAL ───────────────────────────────────────
    st.markdown('<div class="sec">Apuracao PGDAS-D – Credito por Mes</div>', unsafe_allow_html=True)
    seg_ic = {"SEGREGACAO_TOTAL":"✅ TOTAL","SEGREGACAO_PARCIAL":"🟡 PARCIAL",
              "SEM_SEGREGACAO":"🔴 SEM SEG.","SEM_DADO_PGDAS":"– SEM DADO"}
    rows_ap = []
    for r in apuracao:
        l = {"Mes":fmt_mes(r["mes"]),
             "RBT12":brl(r.get("rbt12")),
             "Faixa":str(r.get("faixa","–")),
             "Aliq. Ef.":pct_s(r.get("aliquota_efetiva"),4),
             "Rec. XML":brl(r.get("receita_total"))}
        if usar_pgdas and df_pgdas is not None:
            l["Rec. PGDAS"]  = brl(r.get("receita_pgdas"))
            l["Diverg."]     = ("🔴 " if r.get("alerta_divergencia") else "✅ ") + \
                               pct_s(r.get("divergencia_pct")) if r.get("divergencia_pct") is not None else "–"
            l["Segregacao"]  = seg_ic.get(r.get("status_segregacao","–"),"–")
        l.update({
            "Rec. Monof.":   brl(r.get("receita_monofasica")),
            "Rec. Trib.":    brl(r.get("receita_tributavel")),
            "DAS Usado":     brl(r.get("das_usado")),
            "Fonte":         "🟢 REAL" if r.get("fonte_das")=="REAL" else "🟡 EST.",
            "DAS Correto":   brl(r.get("das_correto")),
            "% PIS+COF":     pct_s(r.get("pct_pis_cofins_val")),
            "Cred. Final":   brl(r.get("credito_final")),
            "Status":        {"RECUPERAVEL":"✅","NAO_RECUPERAVEL":"⚪",
                              "RISCO_FISCAL":"🔴"}.get(r.get("status_credito",""),"–"),
        })
        rows_ap.append(l)
    st.dataframe(pd.DataFrame(rows_ap), use_container_width=True, hide_index=True)

    valid = [r for r in apuracao if r.get("credito_final") is not None]
    t1,t2,t3,t4 = st.columns(4)
    t1.metric("DAS Total Usado",     brl(sum(r.get("das_usado",0) or 0 for r in valid)))
    t2.metric("DAS Total Correto",   brl(sum(r.get("das_correto",0) or 0 for r in valid)))
    t3.metric("Credito Bruto Total", brl(sum(r.get("credito_bruto",0) or 0 for r in valid)))
    t4.metric("Credito Real PIS/COF",brl(sc["credito_total"]),
              delta="a recuperar" if sc["credito_total"] > 0 else None)

    if usar_pgdas and df_pgdas is not None:
        n_r = sum(1 for r in valid if r.get("fonte_das") == "REAL")
        n_e = sum(1 for r in valid if r.get("fonte_das") == "ESTIMADO")
        st.caption("🟢 {} mes(es) com DAS real  |  🟡 {} mes(es) com DAS estimado".format(n_r, n_e))

    # ── OPORTUNIDADE / RISCOS ─────────────────────────────────
    oc, rc = st.columns(2)
    with oc:
        st.markdown('<div class="sec">Oportunidade Identificada</div>', unsafe_allow_html=True)
        top3 = sorted([r for r in apuracao if (r.get("credito_final") or 0) > 0],
                      key=lambda x: x.get("credito_final", 0), reverse=True)[:3]
        for r in top3:
            st.success("**{}** | Credito: {} | Monof.: {}".format(
                fmt_mes(r["mes"]), brl(r.get("credito_final")),
                brl(r.get("receita_monofasica"))))
        if not top3:
            st.info("Nenhum credito recuperavel identificado.")
    with rc:
        st.markdown('<div class="sec">Riscos Tributarios</div>', unsafe_allow_html=True)
        riscos = [i for i in itens if i.get("status_tributario") == "MONOFASICO_COM_RISCO"]
        for i in riscos[:4]:
            st.warning("**{}** | NCM {} | Score {} | {}".format(
                i["descricao"][:30], i["ncm"],
                i.get("score_risco", 0), i.get("motivo_alerta", "")[:50]))
        if not riscos:
            st.success("Nenhum risco tributario identificado.")

    # ── ITENS CLASSIFICADOS ───────────────────────────────────
    st.markdown('<div class="sec">Itens Classificados</div>', unsafe_allow_html=True)
    f1, f2, f3 = st.columns(3)
    filtro_st  = f1.selectbox("Status", ["Todos","MONOFASICO_VALIDADO",
                                          "MONOFASICO_COM_RISCO","NAO_MONOFASICO","INCONSISTENTE"])
    filtro_mes = f2.selectbox("Mes", ["Todos"] + sorted(df_itens["mes"].unique().tolist()))
    filtro_arq = f3.selectbox("Arquivo", ["Todos"] + sorted(df_itens["arquivo"].unique().tolist()))
    df_ex = df_itens.copy()
    if filtro_st  != "Todos": df_ex = df_ex[df_ex["status_tributario"] == filtro_st]
    if filtro_mes != "Todos": df_ex = df_ex[df_ex["mes"] == filtro_mes]
    if filtro_arq != "Todos": df_ex = df_ex[df_ex["arquivo"] == filtro_arq]
    st.dataframe(df_ex[["mes","arquivo","descricao","ncm","cfop","cst_pis","cst_cofins",
                          "valor","status_tributario","score_risco","motivo_alerta"]].rename(columns={
        "mes":"Mes","arquivo":"Arquivo","descricao":"Descricao","ncm":"NCM",
        "cfop":"CFOP","cst_pis":"CST PIS","cst_cofins":"CST COF",
        "valor":"Valor (R$)","status_tributario":"Status",
        "score_risco":"Score","motivo_alerta":"Motivo",
    }), use_container_width=True, height=360)

    # ── RESUMO EXECUTIVO ──────────────────────────────────────
    with st.expander("Resumo Executivo Automatico"):
        st.text(txt)

    # ── EXPORTS ───────────────────────────────────────────────
    st.markdown('<div class="sec">Exportar Relatorios</div>', unsafe_allow_html=True)
    e1, e2, e3 = st.columns(3)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    with e1:
        st.download_button("📥 Excel (6 abas)",
            data=gerar_excel(itens, res, apuracao, sc, alertas, get_logs()),
            file_name="piscofins_{}.xlsx".format(ts),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    with e2:
        st.download_button("📥 Apuracao CSV",
            data=gerar_csv(apuracao),
            file_name="apuracao_{}.csv".format(ts),
            mime="text/csv", use_container_width=True)
    with e3:
        if REPORTLAB_OK:
            st.download_button("📄 PDF Executivo",
                data=gerar_pdf(res, apuracao, sc, txt),
                file_name="relatorio_{}.pdf".format(ts),
                mime="application/pdf", use_container_width=True)
        else:
            st.caption("PDF: `pip install reportlab`")

    # ── LOGS ──────────────────────────────────────────────────
    with st.expander("Logs de Auditoria ({} eventos)".format(len(get_logs()))):
        df_log = pd.DataFrame(get_logs())
        if not df_log.empty:
            ic = {"INFO":"🔵","AVISO":"🟡","ERRO":"🔴","RISCO":"🟠"}
            df_log["nivel"] = df_log["nivel"].apply(lambda x: ic.get(x,"") + x)
            st.dataframe(df_log, use_container_width=True, hide_index=True)

    # ── METODOLOGIA ───────────────────────────────────────────
    with st.expander("Metodologia e Base Legal"):
        st.markdown("""
**Nivel 1 – Triagem:** XMLs classificados por NCM (Tabela 4.3.10 EFD-Contribuicoes) + validacao CST/CFOP.

**Nivel 2 – Apuracao Real:** DAS real do PGDAS substitui DAS estimado. Segregacao validada.

**Segregacao:** TOTAL >= 90% | PARCIAL 5–90% | SEM_SEGREGACAO < 5%.

**Aliquota Efetiva:** `(RBT12 × aliq_nominal − deducao) / RBT12`

**Credito:** `Cred. Bruto = DAS Usado − DAS Correto` | `Cred. Real = Bruto × (% PIS + % COFINS)`

**Base legal:** LC 123/2006 · CGSN 140/2018 · Lei 10.147/2000
        """)

    st.divider()
    st.caption("PIS/COFINS Pro v{} · Analise preliminar · Validar com contador · LC 123/2006".format(VERSAO))


if __name__ == "__main__":
    main()
